import streamlit as st
import pandas as pd
import openpyxl
import math
import io
import zipfile
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ──────────────── UI Configuration ────────────────
st.set_page_config(page_title="Seating Arrangement Generator", layout="wide")
st.title("Seating Arrangement Automation - Internship Project")

st.markdown("""
### Instructions:

**Upload your input Excel with the following sheets:**
--> `in_timetable`
--> `in_course_roll_mapping`
--> `in_roll_name_mapping`
--> `in_room_capacity`

---

**After uploading, choose buffer and density options to download:**
- `overall_seating.xlsx`
- `seats_left.xlsx`
- Individual date/session attendance files zipped inside folders:
    - `date/morning/`
    - `date/evening/`
""")

uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])
if not uploaded_file:
    st.info("Waiting for Excel file upload...")
    st.stop()

# ──────────────── Read Excel Sheets ────────────────
xls = pd.ExcelFile(uploaded_file)
df_timetable = xls.parse("in_timetable")
df_course_roll = xls.parse("in_course_roll_mapping")
df_roll_name = xls.parse("in_roll_name_mapping")
df_rooms = xls.parse("in_room_capacity")

# ──────────────── User Inputs ────────────────
buffer_seats = st.number_input("Buffer seats in each room", min_value=0, value=5)
density_mode = st.radio("Seating Mode", ["Sparse", "Dense"])

# ──────────────── Prepare Mappings ────────────────
df_course_roll.columns = df_course_roll.columns.str.strip()
df_course_roll['course_code'] = df_course_roll['course_code'].str.strip().str.upper()
df_course_roll['rollno'] = df_course_roll['rollno'].str.strip().str.upper()
df_roll_name.columns = df_roll_name.columns.str.strip()
df_roll_name['Roll'] = df_roll_name['Roll'].str.strip().str.upper()

name_lookup = dict(zip(df_roll_name['Roll'], df_roll_name['Name']))
course_to_students = df_course_roll.groupby("course_code")['rollno'].apply(lambda x: sorted(x.tolist())).to_dict()

room_list = []
for _, row in df_rooms.iterrows():
    room_id = str(row['Room No.']).strip()
    capacity = int(row['Exam Capacity'])
    block = str(row['Block']).strip()
    numeric_part = int(room_id) if block == 'B1' else int(room_id.split('-')[-1])
    room_list.append(dict(room=room_id, capacity=capacity, block=block, numeric=numeric_part))

# ──────────────── Room Allocation Logic ────────────────
def assign_students(rolls, available_rooms):
    total_students = len(rolls)
    filtered_rooms = []

    for room in available_rooms:
        effective_capacity = room['capacity'] - buffer_seats
        if effective_capacity <= 0:
            continue
        usable = math.floor(effective_capacity * 0.5) if density_mode == 'Sparse' else effective_capacity
        if usable > 0:
            room.update({'usable': usable})
            filtered_rooms.append(room)

    if sum(r['usable'] for r in filtered_rooms) < total_students:
        return None

    for blk in ['B1', 'B2']:
        block_rooms = sorted([r for r in filtered_rooms if r['block'] == blk], key=lambda x: x['numeric'])
        if sum(r['usable'] for r in block_rooms) >= total_students:
            assignments = []
            index = 0
            for room in block_rooms:
                if index >= total_students:
                    break
                count = min(room['usable'], total_students - index)
                assignments.append((room['room'], rolls[index:index + count]))
                index += count
            return assignments

    sorted_rooms = sorted(filtered_rooms, key=lambda r: -r['usable'])
    output = []
    idx = 0
    for room in sorted_rooms:
        if idx >= total_students:
            break
        assign = min(room['usable'], total_students - idx)
        output.append((room['room'], rolls[idx:idx + assign]))
        idx += assign

    return output if idx >= total_students else None

# ──────────────── Seating Data Creation ────────────────
summary = []
left_out = []
datewise_data = {}

for _, row in df_timetable.iterrows():
    dt_str = pd.to_datetime(row['Date']).strftime("%d_%m_%Y")
    datewise_data.setdefault(dt_str, {'morning': {}, 'evening': {}})

    morning_courses = [] if pd.isna(row['Morning']) else [c.strip() for c in row['Morning'].split(';')]
    evening_courses = [] if pd.isna(row['Evening']) else [c.strip() for c in row['Evening'].split(';')]

    available_rooms = room_list.copy()
    overflow_courses = []

    for course in sorted(morning_courses, key=lambda c: -len(course_to_students.get(c, []))):
        students = course_to_students.get(course, [])
        assignment = assign_students(students, available_rooms)
        if not assignment:
            overflow_courses.append(course)
        else:
            for room_id, grp in assignment:
                available_rooms = [r for r in available_rooms if r['room'] != room_id]
                datewise_data[dt_str]['morning'][(course, room_id)] = grp
                summary.append({'Date': dt_str, 'Course': course, 'Room': room_id, 'Rolls': ";".join(grp)})

    available_rooms = room_list.copy()
    for course in sorted(evening_courses + overflow_courses, key=lambda c: -len(course_to_students.get(c, []))):
        students = course_to_students.get(course, [])
        assignment = assign_students(students, available_rooms)
        if not assignment:
            left_out.append({'Date': dt_str, 'Course': course, 'Unallocated': len(students)})
        else:
            for room_id, grp in assignment:
                available_rooms = [r for r in available_rooms if r['room'] != room_id]
                datewise_data[dt_str]['evening'][(course, room_id)] = grp
                summary.append({'Date': dt_str, 'Course': course, 'Room': room_id, 'Rolls': ";".join(grp)})

# ──────────────── Display Tables ────────────────
st.subheader("Final Seating Plan")
st.dataframe(pd.DataFrame(summary))

st.subheader("Unallocated Seats")
st.dataframe(pd.DataFrame(left_out))

# ──────────────── Excel Styling ────────────────
border_style = Side(style='thin')
def format_worksheet(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(border_style, border_style, border_style, border_style)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    for col_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = length + 2

# ──────────────── ZIP Creation ────────────────
zip_memory = io.BytesIO()
with zipfile.ZipFile(zip_memory, "w") as archive:
    def to_excel_bytes(df, sheet_name):
        bio = io.BytesIO()
        with pd.ExcelWriter(bio, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        return bio.getvalue()

    archive.writestr("overall_seating.xlsx", to_excel_bytes(pd.DataFrame(summary), "Overall"))
    archive.writestr("seats_left.xlsx", to_excel_bytes(pd.DataFrame(left_out), "Left"))

    for dt, sessions in datewise_data.items():
        for session_type in ["morning", "evening"]:
            prefix = f"{dt}/{session_type}/"
            for (course, room), students in sessions[session_type].items():
                wb = Workbook()
                ws = wb.active
                ws.title = session_type

                ws.append([f"Course: {course} | Room: {room} | Date: {dt.replace('_','-')} | Session: {session_type.capitalize()}"])
                ws.append(["Roll", "Student Name", "Signature"])
                for roll in students:
                    ws.append([roll, name_lookup.get(roll, "Unknown"), ""])

                ws.append([])
                for i in range(1, 6):
                    ws.append([f"TA{i}", "", ""])
                ws.append([])
                for i in range(1, 6):
                    ws.append([f"Invigilator{i}", "", ""])

                format_worksheet(ws)
                file_output = io.BytesIO()
                wb.save(file_output)
                archive.writestr(f"{prefix}{dt}_{course}_{room}_{session_type}.xlsx", file_output.getvalue())

zip_memory.seek(0)
st.download_button(
    label="Download ZIP of Seating Plan",
    data=zip_memory,
    file_name="exam_seating.zip",
    mime="application/zip"
)

st.success("All files are ready! You may download them using the button above.")
